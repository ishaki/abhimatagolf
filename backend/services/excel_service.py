from typing import List, Dict, Any, Optional
import pandas as pd
from io import BytesIO
from sqlmodel import Session, select
from models.participant import Participant
from models.scorecard import Scorecard
from models.event import Event
from models.event_division import EventDivision
from models.course import Course, Hole
from schemas.participant import ParticipantResponse
from schemas.scorecard import ScorecardResponse
from core.app_logging import logger
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet


class ExcelService:
    def __init__(self, session: Session):
        self.session = session

    def export_participants_to_excel(self, event_id: int) -> BytesIO:
        """Export participants for an event to Excel format."""
        try:
            # Get event information
            event = self.session.get(Event, event_id)
            if not event:
                raise ValueError(f"Event with ID {event_id} not found")

            # Get participants with their division information
            participants = self.session.exec(
                select(Participant, EventDivision)
                .join(EventDivision, Participant.division_id == EventDivision.id, isouter=True)
                .where(Participant.event_id == event_id)
            ).all()

            # Prepare data for Excel
            data = []
            for participant, division in participants:
                data.append({
                    'Name': participant.name,
                    'Handicap': participant.declared_handicap,
                    'Division': division.name if division else participant.division or 'N/A',
                    'Division ID': participant.division_id or '',
                    'Country': participant.country or '',
                    'Sex': participant.sex or '',
                    'Phone No': participant.phone_no or '',
                    'Event Status': participant.event_status or 'Ok',
                    'Event Description': participant.event_description or '',
                    'Registered At': participant.registered_at.strftime('%Y-%m-%d %H:%M') if participant.registered_at else '',
                })

            # Create DataFrame
            df = pd.DataFrame(data)

            # Create Excel file in memory
            output = BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Write participants data
                df.to_excel(writer, sheet_name='Participants', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Participants']
                
                # Apply formatting
                self._format_participants_sheet(worksheet, event.name)
                
                # Add summary sheet
                self._add_summary_sheet(workbook, event, participants)

            output.seek(0)
            logger.info(f"Successfully exported {len(participants)} participants for event {event_id}")
            return output

        except Exception as e:
            logger.error(f"Error exporting participants to Excel: {str(e)}")
            raise

    def export_participant_scores_detailed(self, event_id: int) -> BytesIO:
        """
        Export detailed participant scores to Excel with three sheets:
        1. Scores: Participant, Hole 1-9, Total Out, Hole 10-18, Total In, Total (showing scores)
        2. Points: Participant, Hole 1-9, Total Out, Hole 10-18, Total In, Total Point (showing System 36 points)
        3. Summary: Participant, Declared Hcp, Course Handicap, Total Gross, Nett, Total Point
        
        Note: For System 36 events, the Course Handicap column shows System 36 Handicap (36 - Total Points),
        and Nett is calculated as Total Gross - System 36 Handicap.
        For other events, Nett = Total Gross - Course Handicap.
        """
        try:
            from models.event import Event, ScoringType, System36Variant
            from models.participant import Participant
            from models.scorecard import Scorecard
            from models.course import Hole
            from models.event_division import EventDivision, DivisionType
            from services.scoring_strategies.system36 import System36ScoringStrategy
            
            # Get event
            event = self.session.get(Event, event_id)
            if not event:
                raise ValueError(f"Event {event_id} not found")
            
            # Get all participants for the event
            participants_query = select(Participant).where(Participant.event_id == event_id)
            participants = list(self.session.exec(participants_query).all())
            
            if not participants:
                raise ValueError(f"No participants found for event {event_id}")
            
            # Get all holes for the course
            holes_query = select(Hole).where(Hole.course_id == event.course_id).order_by(Hole.number)
            holes = list(self.session.exec(holes_query).all())
            
            # Get divisions for participant lookup
            divisions_query = select(EventDivision).where(EventDivision.event_id == event_id)
            divisions = list(self.session.exec(divisions_query).all())
            division_map = {div.id: div for div in divisions}

            # Get winner results to check for division reassignments
            from models.winner_result import WinnerResult
            winner_results_query = select(WinnerResult).where(WinnerResult.event_id == event_id)
            winner_results = list(self.session.exec(winner_results_query).all())
            winner_result_map = {wr.participant_id: wr for wr in winner_results}
            
            # Create Excel file in memory
            output = BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Prepare data for all three sheets
                scores_data = []
                points_data = []
                summary_data = []
                
                # Initialize System 36 strategy if needed
                system36_strategy = None
                if event.scoring_type == ScoringType.SYSTEM_36:
                    system36_strategy = System36ScoringStrategy()
                
                for participant in participants:
                    # Get all scorecards for this participant
                    scorecards_query = select(Scorecard).where(Scorecard.participant_id == participant.id)
                    scorecards = list(self.session.exec(scorecards_query).all())
                    
                    # Create mapping of hole_id to scorecard
                    scorecard_map = {sc.hole_id: sc for sc in scorecards}
                    
                    # Get division name
                    division_name = participant.division
                    if participant.division_id and participant.division_id in division_map:
                        division_name = division_map[participant.division_id].name
                    
                    # Initialize row data for Sheet 1 (Scores)
                    scores_row = {
                        'Participant': participant.name
                    }
                    
                    # Initialize row data for Sheet 2 (Points)
                    points_row = {
                        'Participant': participant.name
                    }
                    
                    # Process each hole and calculate totals
                    total_gross = 0
                    total_points = 0
                    out_total = 0  # Front 9 (holes 1-9)
                    in_total = 0   # Back 9 (holes 10-18)
                    out_points = 0
                    in_points = 0
                    
                    # First pass: calculate totals
                    for hole in holes:
                        scorecard = scorecard_map.get(hole.id)
                        strokes = scorecard.strokes if scorecard and scorecard.strokes > 0 else 0
                        
                        if strokes > 0:
                            total_gross += strokes
                            # Add to front/back nine scores
                            if hole.number <= 9:
                                out_total += strokes
                            else:
                                in_total += strokes
                        
                        # Calculate points for System 36
                        if event.scoring_type == ScoringType.SYSTEM_36 and system36_strategy and strokes > 0:
                            points = system36_strategy.calculate_system36_points(strokes, hole.par)
                            total_points += points
                            
                            # Add to front/back nine points
                            if hole.number <= 9:
                                out_points += points
                            else:
                                in_points += points
                    
                    # Second pass: build rows with proper column order for Sheet 1 (Scores)
                    # Front 9 holes (1-9) - Column names: "Hole 1" through "Hole 9"
                    for hole in holes:
                        if hole.number <= 9:
                            scorecard = scorecard_map.get(hole.id)
                            strokes = scorecard.strokes if scorecard and scorecard.strokes > 0 else 0
                            scores_row[f'Hole {hole.number}'] = strokes if strokes > 0 else ''
                    
                    # Out subtotal
                    scores_row['Total Out'] = out_total if out_total > 0 else ''
                    
                    # Back 9 holes (10-18) - Column names: "Hole 10" through "Hole 18"
                    for hole in holes:
                        if hole.number > 9:
                            scorecard = scorecard_map.get(hole.id)
                            strokes = scorecard.strokes if scorecard and scorecard.strokes > 0 else 0
                            scores_row[f'Hole {hole.number}'] = strokes if strokes > 0 else ''
                    
                    # In subtotal
                    scores_row['Total In'] = in_total if in_total > 0 else ''
                    
                    # Total
                    scores_row['Total'] = total_gross if total_gross > 0 else ''
                    
                    # Build rows for Sheet 2 (Points) - System 36 only
                    if event.scoring_type == ScoringType.SYSTEM_36:
                        # Front 9 holes points (1-9) - Column names: "Hole 1" through "Hole 9"
                        for hole in holes:
                            if hole.number <= 9:
                                scorecard = scorecard_map.get(hole.id)
                                strokes = scorecard.strokes if scorecard and scorecard.strokes > 0 else 0
                                if strokes > 0 and system36_strategy:
                                    points = system36_strategy.calculate_system36_points(strokes, hole.par)
                                    points_row[f'Hole {hole.number}'] = points
                                else:
                                    points_row[f'Hole {hole.number}'] = ''
                        
                        # Out subtotal
                        points_row['Total Out'] = out_points if out_total > 0 else ''
                        
                        # Back 9 holes points (10-18) - Column names: "Hole 10" through "Hole 18"
                        for hole in holes:
                            if hole.number > 9:
                                scorecard = scorecard_map.get(hole.id)
                                strokes = scorecard.strokes if scorecard and scorecard.strokes > 0 else 0
                                if strokes > 0 and system36_strategy:
                                    points = system36_strategy.calculate_system36_points(strokes, hole.par)
                                    points_row[f'Hole {hole.number}'] = points
                                else:
                                    points_row[f'Hole {hole.number}'] = ''
                        
                        # In subtotal
                        points_row['Total In'] = in_points if in_total > 0 else ''
                        
                        # Total Point
                        points_row['Total Point'] = total_points if total_gross > 0 else ''
                    else:
                        # For non-System 36 events, still create the structure but leave empty
                        for hole in holes:
                            if hole.number <= 9:
                                points_row[f'Hole {hole.number}'] = ''
                        points_row['Total Out'] = ''
                        for hole in holes:
                            if hole.number > 9:
                                points_row[f'Hole {hole.number}'] = ''
                        points_row['Total In'] = ''
                        points_row['Total Point'] = ''
                    
                    # Calculate Nett score based on scoring type
                    if event.scoring_type == ScoringType.SYSTEM_36:
                        # For System 36: Nett = Total Gross - System 36 Handicap
                        # System 36 Handicap = 36 - Total Points (only if 18 holes completed)
                        holes_completed = sum(1 for hole in holes if scorecard_map.get(hole.id) and scorecard_map.get(hole.id).strokes > 0)
                        if holes_completed >= 18:
                            system36_handicap = 36 - total_points
                            nett_score = total_gross - system36_handicap if total_gross > 0 else ''
                            course_handicap = system36_handicap
                        else:
                            # Incomplete round - can't calculate System 36 handicap
                            course_handicap = participant.course_handicap
                            nett_score = ''
                    else:
                        # For other scoring types: Nett = Total Gross - Course Handicap
                        course_handicap = participant.course_handicap
                        nett_score = total_gross - course_handicap if total_gross > 0 else ''
                    
                    # Get current division name
                    current_division_name = division_name or participant.division or ''

                    # Check for division reassignment from winner results
                    original_division_name = current_division_name  # Default to current
                    updated_division_name = current_division_name   # Default to current

                    winner_result = winner_result_map.get(participant.id)
                    if winner_result and winner_result.original_division_id:
                        # This participant was reassigned - get original division name
                        original_division = division_map.get(winner_result.original_division_id)
                        if original_division:
                            original_division_name = original_division.name
                            updated_division_name = current_division_name

                    # Check for disqualification (System 36 Modified validation)
                    is_disqualified = False
                    disqualification_reason = ''

                    # Special award winners are EXEMPT from disqualification
                    # (They compete cross-division for best gross/net, not within division)
                    is_special_award_winner = winner_result and winner_result.award_category is not None

                    if (not is_special_award_winner and
                        event.scoring_type == ScoringType.SYSTEM_36 and
                        event.system36_variant == System36Variant.MODIFIED and
                        total_gross > 0):

                        # Get calculated System 36 handicap
                        holes_completed = sum(1 for hole in holes if scorecard_map.get(hole.id) and scorecard_map.get(hole.id).strokes > 0)

                        if holes_completed >= 18 and participant.division_id:
                            calculated_handicap = 36 - total_points
                            current_division = division_map.get(participant.division_id)

                            # Only validate for Men's divisions with defined handicap ranges
                            if current_division and current_division.division_type == DivisionType.MEN:
                                current_min = current_division.handicap_min
                                current_max = current_division.handicap_max

                                if current_min is not None and current_max is not None:
                                    # Check if calculated handicap exceeds division maximum
                                    if calculated_handicap > current_max:
                                        is_disqualified = True
                                        disqualification_reason = f'System 36 HCP ({calculated_handicap:.1f}) > Division Max ({current_max})'
                                    # Check if calculated handicap is below division minimum (and no reassignment found)
                                    elif calculated_handicap < current_min and not winner_result:
                                        is_disqualified = True
                                        disqualification_reason = f'System 36 HCP ({calculated_handicap:.1f}) < Division Min ({current_min}), no appropriate division'

                    # Build notes column (includes special awards, division winners, and disqualification reasons)
                    notes = ''
                    if is_special_award_winner and winner_result:
                        # Show special award category (e.g., "Best Gross Winner", "Best Net Winner")
                        notes = f'{winner_result.award_category} Winner'
                    elif winner_result and winner_result.division_rank:
                        # Show division/group winner ranking with division name
                        rank_suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(winner_result.division_rank, 'th')
                        division_name = winner_result.division or current_division_name or 'Division'
                        notes = f'{division_name} - {winner_result.division_rank}{rank_suffix} Place'
                    elif is_disqualified:
                        # Show disqualification reason
                        notes = disqualification_reason

                    # Build row for Sheet 3 (Summary) - ordered columns
                    summary_row_ordered = {
                        'Participant': participant.name,
                        'Original Division': original_division_name,
                        'Updated Division': updated_division_name if updated_division_name != original_division_name else original_division_name,
                        'Declared Hcp': participant.declared_handicap,
                        'Course Handicap': course_handicap if total_gross > 0 else '',
                        'Total Gross': total_gross if total_gross > 0 else '',
                        'Nett': nett_score if isinstance(nett_score, int) or isinstance(nett_score, float) else '',
                        'Total Point': total_points if event.scoring_type == ScoringType.SYSTEM_36 and total_gross > 0 else '',
                        'Disqualified': 'Yes' if is_disqualified else 'No',
                        'Notes': notes
                    }
                    
                    scores_data.append(scores_row)
                    points_data.append(points_row)
                    summary_data.append(summary_row_ordered)
                
                # Create Sheet 1: Scores
                scores_df = pd.DataFrame(scores_data)
                scores_df.to_excel(writer, sheet_name='Scores', index=False)
                
                # Create Sheet 2: Points
                points_df = pd.DataFrame(points_data)
                points_df.to_excel(writer, sheet_name='Points', index=False)
                
                # Create Sheet 3: Summary
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Add note to Sheet 3 (Summary) about Course Handicap calculation
                summary_worksheet = writer.sheets['Summary']
                max_row = len(summary_data) + 1  # +1 for header row
                
                # Add empty row
                summary_worksheet[f'A{max_row + 2}'] = "Note:"
                note_cell = summary_worksheet[f'A{max_row + 2}']
                note_cell.font = Font(bold=True, color="0000FF")
                
                # Add note text explaining handicap calculation (depends on scoring type)
                if event.scoring_type == ScoringType.SYSTEM_36:
                    note_text = "System 36 Handicap = 36 - Total Points (calculated from your performance)"
                    explanation_text = "In System 36, your handicap is derived from your total points. Lower points result in higher handicap, and vice versa."
                else:
                    note_text = "Course Handicap = (Declared Handicap × Teebox Slope Rating) / 113"
                    explanation_text = "Course Handicap adjusts your declared handicap based on the difficulty of the teebox you're playing from."
                
                summary_worksheet[f'A{max_row + 3}'] = note_text
                note_text_cell = summary_worksheet[f'A{max_row + 3}']
                note_text_cell.font = Font(italic=True, color="666666")
                
                summary_worksheet[f'A{max_row + 4}'] = explanation_text
                explanation_cell = summary_worksheet[f'A{max_row + 4}']
                explanation_cell.font = Font(italic=True, color="666666")
                
                # Add empty row before Nett calculation note
                summary_worksheet[f'A{max_row + 6}'] = "Nett Calculation:"
                nett_note_cell = summary_worksheet[f'A{max_row + 6}']
                nett_note_cell.font = Font(bold=True, color="0000FF")
                
                # Add note text explaining Nett calculation
                if event.scoring_type == ScoringType.SYSTEM_36:
                    nett_formula = "Nett = Total Gross - System 36 Handicap (where System 36 Handicap = 36 - Total Points)"
                    nett_explanation_text = "In System 36, your handicap is calculated from your total points (36 - Total Points). Nett shows your adjusted performance."
                else:
                    nett_formula = "Nett = Total Gross - Course Handicap"
                    nett_explanation_text = "Nett score is your gross score minus your course handicap, showing your adjusted performance level."
                
                summary_worksheet[f'A{max_row + 7}'] = nett_formula
                nett_formula_cell = summary_worksheet[f'A{max_row + 7}']
                nett_formula_cell.font = Font(italic=True, color="666666")
                
                summary_worksheet[f'A{max_row + 8}'] = nett_explanation_text
                nett_explanation_cell = summary_worksheet[f'A{max_row + 8}']
                nett_explanation_cell.font = Font(italic=True, color="666666")
                
                # Auto-adjust column widths
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 20)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
            
            output.seek(0)
            return output
            
        except Exception as e:
            logger.error(f"Error in export_participant_scores_detailed: {str(e)}")
            raise

    def export_scorecards_to_excel(self, event_id: int) -> BytesIO:
        """Export scorecards for an event to Excel format."""
        try:
            # Get event information
            event = self.session.get(Event, event_id)
            if not event:
                raise ValueError(f"Event with ID {event_id} not found")

            # Get course information
            course = self.session.get(Course, event.course_id)
            if not course:
                raise ValueError(f"Course not found for event {event_id}")

            # Get holes for the course
            holes = self.session.exec(
                select(Hole).where(Hole.course_id == course.id).order_by(Hole.hole_number)
            ).all()

            # Get participants with their scorecards
            participants = self.session.exec(
                select(Participant, EventDivision)
                .join(EventDivision, Participant.division_id == EventDivision.id, isouter=True)
                .where(Participant.event_id == event_id)
            ).all()

            # Create Excel file in memory
            output = BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Create scorecard data
                scorecard_data = []
                
                for participant, division in participants:
                    # Get scorecards for this participant
                    scorecards = self.session.exec(
                        select(Scorecard, Hole)
                        .join(Hole, Scorecard.hole_id == Hole.id)
                        .where(Scorecard.participant_id == participant.id)
                        .order_by(Hole.hole_number)
                    ).all()

                    # Create a row for this participant
                    row = {
                        'Player': participant.name,
                        'Division': division.name if division else participant.division or 'N/A',
                        'Handicap': participant.declared_handicap,
                    }

                    # Add hole scores
                    hole_scores = {}
                    total_gross = 0
                    total_par = 0
                    
                    for scorecard, hole in scorecards:
                        hole_scores[hole.hole_number] = scorecard.score
                        if scorecard.score is not None:
                            total_gross += scorecard.score
                            total_par += hole.par

                    # Add hole columns
                    for hole in holes:
                        row[f'Hole {hole.hole_number}'] = hole_scores.get(hole.hole_number, '')
                        row[f'Par {hole.hole_number}'] = hole.par

                    # Add totals
                    row['Total Gross'] = total_gross if total_gross > 0 else ''
                    row['Total Par'] = total_par
                    row['Net Score'] = total_gross - participant.declared_handicap if total_gross > 0 else ''
                    row['To Par'] = total_gross - total_par if total_gross > 0 else ''

                    scorecard_data.append(row)

                # Create DataFrame
                df = pd.DataFrame(scorecard_data)

                # Write to Excel
                df.to_excel(writer, sheet_name='Scorecards', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Scorecards']
                
                # Apply formatting
                self._format_scorecards_sheet(worksheet, event.name, course.name)
                
                # Add summary sheet
                self._add_scorecard_summary_sheet(workbook, event, participants, holes)

            output.seek(0)
            logger.info(f"Successfully exported scorecards for {len(participants)} participants for event {event_id}")
            return output

        except Exception as e:
            logger.error(f"Error exporting scorecards to Excel: {str(e)}")
            raise

    def generate_participant_template(self, event_id: Optional[int] = None) -> BytesIO:
        """Generate Excel template for participant upload."""
        try:
            # Create template data with all fields
            template_data = [
                {
                    'name': 'John Doe',
                    'declared_handicap': 12,
                    'division': 'Championship',
                    'division_id': 1,
                    'country': 'United States',
                    'sex': 'Male',
                    'phone_no': '+1234567890',
                    'event_status': 'Ok',
                    'event_description': 'Regular participant'
                },
                {
                    'name': 'Jane Smith',
                    'declared_handicap': 8,
                    'division': 'Ladies',
                    'division_id': 2,
                    'country': 'United Kingdom',
                    'sex': 'Female',
                    'phone_no': '+44123456789',
                    'event_status': 'Ok',
                    'event_description': ''
                },
                {
                    'name': 'Bob Johnson',
                    'declared_handicap': 18,
                    'division': 'Senior',
                    'division_id': 3,
                    'country': 'Canada',
                    'sex': 'Male',
                    'phone_no': '+1987654321',
                    'event_status': 'Ok',
                    'event_description': ''
                }
            ]

            df = pd.DataFrame(template_data)

            # Get event divisions if event_id is provided
            divisions_list = []
            if event_id:
                divisions = self.session.exec(
                    select(EventDivision).where(EventDivision.event_id == event_id)
                ).all()
                divisions_list = [(div.id, div.name, div.handicap_min, div.handicap_max) for div in divisions]

            # Create Excel file in memory
            output = BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Write template data
                df.to_excel(writer, sheet_name='Participants', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Participants']
                
                # Apply formatting
                self._format_template_sheet(worksheet)
                
                # Add instructions sheet
                self._add_instructions_sheet(workbook)
                
                # Add divisions reference sheet if divisions exist
                if divisions_list:
                    self._add_divisions_sheet(workbook, divisions_list)

            output.seek(0)
            logger.info("Successfully generated participant template")
            return output

        except Exception as e:
            logger.error(f"Error generating participant template: {str(e)}")
            raise

    def _format_participants_sheet(self, worksheet: Worksheet, event_name: str):
        """Apply formatting to participants sheet."""
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Add title
        worksheet.insert_rows(1)
        worksheet.merge_cells('A1:E1')
        title_cell = worksheet['A1']
        title_cell.value = f"Participants - {event_name}"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center")

    def _format_scorecards_sheet(self, worksheet: Worksheet, event_name: str, course_name: str):
        """Apply formatting to scorecards sheet."""
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 15)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Add title
        worksheet.insert_rows(1)
        worksheet.merge_cells('A1:Z1')
        title_cell = worksheet['A1']
        title_cell.value = f"Scorecards - {event_name} ({course_name})"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center")

    def _format_template_sheet(self, worksheet: Worksheet):
        """Apply formatting to template sheet."""
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Add title
        worksheet.insert_rows(1)
        worksheet.merge_cells('A1:I1')
        title_cell = worksheet['A1']
        title_cell.value = "Participant Upload Template"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center")

    def _add_summary_sheet(self, workbook, event: Event, participants: List):
        """Add summary sheet to participants export."""
        ws = workbook.create_sheet("Summary")
        
        # Add summary data
        summary_data = [
            ["Event Name", event.name],
            ["Event Date", event.start_date.strftime('%Y-%m-%d') if event.start_date else 'N/A'],
            ["Course", event.course.name if event.course else 'N/A'],
            ["Total Participants", len(participants)],
            ["Export Date", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ]
        
        for i, (label, value) in enumerate(summary_data, 1):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _add_scorecard_summary_sheet(self, workbook, event: Event, participants: List, holes: List):
        """Add summary sheet to scorecards export."""
        ws = workbook.create_sheet("Summary")
        
        # Add summary data
        summary_data = [
            ["Event Name", event.name],
            ["Event Date", event.start_date.strftime('%Y-%m-%d') if event.start_date else 'N/A'],
            ["Course", event.course.name if event.course else 'N/A'],
            ["Total Participants", len(participants)],
            ["Total Holes", len(holes)],
            ["Export Date", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ]
        
        for i, (label, value) in enumerate(summary_data, 1):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _add_instructions_sheet(self, workbook):
        """Add instructions sheet to template."""
        ws = workbook.create_sheet("Instructions")
        
        instructions = [
            "Participant Upload Template Instructions",
            "",
            "Required Columns:",
            "• name: Participant's full name (required)",
            "",
            "Optional Columns:",
            "• declared_handicap: Golf handicap (0-54, default: 0)",
            "• division: Division name (optional)",
            "• division_id: Division ID from event divisions (optional, see Divisions sheet)",
            "• country: Country name (optional)",
            "• sex: Male or Female (optional)",
            "• phone_no: Phone number with country code (optional, e.g., +1234567890)",
            "• event_status: Ok (default), No Show, or Disqualified",
            "• event_description: Additional notes about participant (optional)",
            "",
            "Instructions:",
            "1. Fill in participant information in the 'Participants' sheet",
            "2. Remove example rows before uploading",
            "3. Ensure names are unique within the event",
            "4. Handicap values must be between 0 and 54",
            "5. Division ID must reference an existing event division (see Divisions sheet)",
            "6. Sex must be exactly 'Male' or 'Female'",
            "7. Event status must be 'Ok', 'No Show', or 'Disqualified'",
            "8. Phone numbers can contain numbers, +, spaces, hyphens, and parentheses",
            "",
            "File Format:",
            "• Save as Excel (.xlsx) or CSV (.csv) format",
            "• Do not modify column headers",
            "• Leave optional fields empty if not needed",
            "",
            "Validation:",
            "• Names cannot be empty",
            "• Handicaps must be numeric (0-54)",
            "• Division IDs must be valid integers",
            "• Sex must be Male or Female (case-sensitive)",
            "• Event status must be Ok, No Show, or Disqualified",
            "",
            "For support, contact the tournament administrator."
        ]
        
        for i, instruction in enumerate(instructions, 1):
            ws[f'A{i}'] = instruction
            if i == 1:  # Title
                ws[f'A{i}'].font = Font(bold=True, size=14)
            elif instruction.startswith("•"):  # Bullet points
                ws[f'A{i}'].font = Font(bold=True)
        
        # Auto-adjust column width
        ws.column_dimensions['A'].width = 80
    
    def _add_divisions_sheet(self, workbook, divisions: List):
        """Add divisions reference sheet to template."""
        ws = workbook.create_sheet("Divisions")
        
        # Add header
        headers = ["Division ID", "Division Name", "Min Handicap", "Max Handicap"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add divisions data
        for row_idx, (div_id, div_name, min_hcp, max_hcp) in enumerate(divisions, 2):
            ws.cell(row=row_idx, column=1, value=div_id)
            ws.cell(row=row_idx, column=2, value=div_name)
            ws.cell(row=row_idx, column=3, value=min_hcp if min_hcp is not None else 'N/A')
            ws.cell(row=row_idx, column=4, value=max_hcp if max_hcp is not None else 'N/A')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add title
        ws.insert_rows(1)
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = "Event Divisions Reference"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
